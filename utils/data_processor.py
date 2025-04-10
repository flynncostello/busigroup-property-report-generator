"""
Data processor module for extracting and analyzing property data from Excel/CSV files.

This module handles all data processing tasks required to generate the property report.
"""

import os
import logging
import pandas as pd
import numpy as np
import sys
import base64
import io
from PIL import Image
import openpyxl
from openpyxl_image_loader import SheetImageLoader

# Set up logger for this module
logger = logging.getLogger(__name__)

def process_excel_data(df, excel_file_path=None):
    """
    Process the Excel/CSV data and extract relevant information for the report.
    
    Args:
        df (pandas.DataFrame): The dataframe containing property data
        excel_file_path (str, optional): Path to the original Excel file for direct image extraction
        
    Returns:
        dict: A dictionary containing all processed data needed for the report
    """
    logger.info("Starting data processing")
    logger.info(f"DataFrame shape: {df.shape}")
    logger.info(f"DataFrame columns: {list(df.columns)}")
    
    # Initialize dictionary to hold images extracted from Excel
    image_dict = {}
    
    # Extract images directly from Excel file if path is provided
    if excel_file_path and os.path.exists(excel_file_path) and excel_file_path.endswith(('.xlsx', '.xls')):
        try:
            logger.info(f"Loading Excel file for image extraction: {excel_file_path}")
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            image_loader = SheetImageLoader(sheet)
            
            # Find the Property Photo column index in Excel
            property_photo_col = None
            for col_idx, header in enumerate(sheet[1], 1):  # Assuming headers are in row 1
                if header.value == "Property Photo":
                    property_photo_col = col_idx
                    logger.info(f"Found 'Property Photo' column at index {property_photo_col}")
                    break
            
            if property_photo_col:
                # Iterate through each row, extract image if present in Property Photo column
                for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skipping header)
                    cell_coord = f"{openpyxl.utils.get_column_letter(property_photo_col)}{row_idx}"
                    
                    # Check if cell has an image
                    if image_loader.image_in(cell_coord):
                        logger.info(f"Found image in cell {cell_coord}")
                        try:
                            # Get the image
                            image = image_loader.get(cell_coord)
                            
                            # Convert image to data URL
                            img_byte_arr = io.BytesIO()
                            image.save(img_byte_arr, format='JPEG')
                            img_byte_arr.seek(0)
                            img_bytes = img_byte_arr.getvalue()
                            b64_image = base64.b64encode(img_bytes).decode('utf-8')
                            data_url = f"data:image/jpeg;base64,{b64_image}"
                            
                            # Store in dictionary with row index as key
                            image_dict[row_idx] = data_url
                            logger.info(f"Successfully extracted image from cell {cell_coord} (size: {len(img_bytes)} bytes)")
                            logger.info(f"Data URL starts with: {data_url[:30]}...")
                        except Exception as e:
                            logger.error(f"Error extracting image from cell {cell_coord}: {str(e)}")
                
                logger.info(f"Extracted {len(image_dict)} images from Excel file")
            else:
                logger.warning("Could not find 'Property Photo' column in Excel sheet")
        except ImportError:
            logger.error("openpyxl or openpyxl_image_loader not installed. Cannot extract images.")
        except Exception as e:
            logger.error(f"Error loading Excel file for image extraction: {str(e)}")
    else:
        logger.warning(f"Excel file path not provided or invalid: {excel_file_path}")
    
    # Validate required columns
    required_columns = [
        'Type', 'Property Photo', 'Street Address', 'Suburb', 'State', 'Postcode',
        'Site Zoning', 'Property Type', 'Car', 'Floor Size (m²)', 'Sale Price',
        'Last Listed Price', 'Last Rental Price', 'PUT IN REPORT (T/F)', 'Busi\'s Comment', '$/m²'
    ]
    
    # Check for missing columns
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logger.error(f"Missing required columns: {missing_columns}")
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Initialize the result dictionary
    result = {
        'for_lease_properties': [],
        'for_sale_properties': [],
        'statistics': {},
    }
    
    # Process statistics for the map page
    try:
        logger.info("Processing statistics for the map page")
        
        # Count properties by type
        for_lease_count = len(df[df['Type'] == 'For Lease'])
        already_leased_count = len(df[df['Type'] == 'Already Leased'])
        for_sale_count = len(df[df['Type'] == 'For Sale'])
        sold_count = len(df[df['Type'] == 'Sold'])
        
        logger.info(f"Property counts - For Lease: {for_lease_count}, Already Leased: {already_leased_count}, "
                   f"For Sale: {for_sale_count}, Sold: {sold_count}")
        
        # Count properties that meet criteria (PUT IN REPORT = T) by type
        for_lease_criteria = len(df[(df['Type'] == 'For Lease') & (df['PUT IN REPORT (T/F)'] == 'T')])
        already_leased_criteria = len(df[(df['Type'] == 'Already Leased') & (df['PUT IN REPORT (T/F)'] == 'T')])
        for_sale_criteria = len(df[(df['Type'] == 'For Sale') & (df['PUT IN REPORT (T/F)'] == 'T')])
        sold_criteria = len(df[(df['Type'] == 'Sold') & (df['PUT IN REPORT (T/F)'] == 'T')])
        
        logger.info(f"Properties meeting criteria - For Lease: {for_lease_criteria}, Already Leased: {already_leased_criteria}, "
                   f"For Sale: {for_sale_criteria}, Sold: {sold_criteria}")
        
        # Calculate average $/m² for each type that meets criteria
        for_lease_avg_price = calculate_average_price_per_sqm(df, 'For Lease', 'T')
        already_leased_avg_price = calculate_average_price_per_sqm(df, 'Already Leased', 'T')
        for_sale_avg_price = calculate_average_price_per_sqm(df, 'For Sale', 'T')
        sold_avg_price = calculate_average_price_per_sqm(df, 'Sold', 'T')
        
        logger.info(f"Average prices $/m² - For Lease: ${for_lease_avg_price}, Already Leased: ${already_leased_avg_price}, "
                   f"For Sale: ${for_sale_avg_price}, Sold: ${sold_avg_price}")
        
        # Store statistics in the result
        result['statistics'] = {
            'for_lease': {
                'total': for_lease_count,
                'criteria': for_lease_criteria,
                'avg_price': for_lease_avg_price
            },
            'already_leased': {
                'total': already_leased_count,
                'criteria': already_leased_criteria,
                'avg_price': already_leased_avg_price
            },
            'for_sale': {
                'total': for_sale_count,
                'criteria': for_sale_criteria,
                'avg_price': for_sale_avg_price
            },
            'sold': {
                'total': sold_count,
                'criteria': sold_criteria,
                'avg_price': sold_avg_price
            }
        }
        
        logger.info("Statistics processed successfully")
    except Exception as e:
        logger.error(f"Error processing statistics: {str(e)}", exc_info=True)
        raise
    
    # Extract 'For Lease' properties to include in the report
    try:
        logger.info("Processing 'For Lease' properties")
        lease_properties = df[(df['Type'] == 'For Lease') & (df['PUT IN REPORT (T/F)'] == 'T')]
        
        for idx, property_row in lease_properties.iterrows():
            # Get image data from the extracted images dictionary if available
            image_data = image_dict.get(idx + 2)  # +2 because Excel rows start from 1 and row 1 is header
            property_data = extract_property_data(property_row, 'For Lease', image_data)
            result['for_lease_properties'].append(property_data)
            
        logger.info(f"Processed {len(result['for_lease_properties'])} 'For Lease' properties")
    except Exception as e:
        logger.error(f"Error processing 'For Lease' properties: {str(e)}", exc_info=True)
        raise
    
    # Extract 'For Sale' properties to include in the report
    try:
        logger.info("Processing 'For Sale' properties")
        sale_properties = df[(df['Type'] == 'For Sale') & (df['PUT IN REPORT (T/F)'] == 'T')]
        
        for idx, property_row in sale_properties.iterrows():
            # Get image data from the extracted images dictionary if available
            image_data = image_dict.get(idx + 2)  # +2 because Excel rows start from 1 and row 1 is header
            property_data = extract_property_data(property_row, 'For Sale', image_data)
            result['for_sale_properties'].append(property_data)
            
        logger.info(f"Processed {len(result['for_sale_properties'])} 'For Sale' properties")
    except Exception as e:
        logger.error(f"Error processing 'For Sale' properties: {str(e)}", exc_info=True)
        raise
    
    return result

def calculate_average_price_per_sqm(df, property_type, put_in_report):
    """
    Calculate the average price per square meter for properties of a given type.
    
    Args:
        df (pandas.DataFrame): The dataframe containing property data
        property_type (str): The type of property ('For Lease', 'Already Leased', 'For Sale', 'Sold')
        put_in_report (str): Filter for 'PUT IN REPORT (T/F)' column ('T' or 'F')
        
    Returns:
        int: The average price per square meter as an integer (rounded)
    """
    # Create a proper copy to avoid SettingWithCopyWarning
    filtered_df = df[(df['Type'] == property_type) & (df['PUT IN REPORT (T/F)'] == put_in_report)].copy()
    
    logger.debug(f"Calculating average $/m² for {property_type} (PUT IN REPORT = {put_in_report})")
    logger.debug(f"Found {len(filtered_df)} matching properties")
    
    if filtered_df.empty:
        logger.debug(f"No properties found for {property_type} with PUT IN REPORT = {put_in_report}")
        return 0
    
    # Convert '$/m²' column to numeric, handling non-numeric values
    if '$/m²' in filtered_df.columns:
        # Create a temporary series from the column to avoid FutureWarning
        # and perform operations on it instead of directly on the DataFrame
        temp_series = pd.Series(filtered_df['$/m²']).astype(str)
        temp_series = temp_series.str.replace('$', '', regex=False)
        temp_series = temp_series.str.replace(',', '', regex=False)
        numeric_values = pd.to_numeric(temp_series, errors='coerce')
        
        # Assign back to DataFrame
        filtered_df.loc[:, '$/m²'] = numeric_values
        
        # Calculate average, ignoring NaN values
        avg_price = numeric_values.mean()
        return round(avg_price) if not np.isnan(avg_price) else 0
    else:
        logger.warning("$/m² column not found in filtered dataframe")
        return 0

def extract_property_data(row, property_type, image_data=None):
    """
    Extract and format property data from a row in the dataframe.
    
    Args:
        row (pandas.Series): A row from the properties dataframe
        property_type (str): The type of property ('For Lease' or 'For Sale')
        image_data (str, optional): Base64 image data if available
        
    Returns:
        dict: A dictionary with the formatted property data
    """
    logger.debug(f"Extracting property data for {property_type} - Suburb: {row['Suburb']}")
    
    # Format street address (capitalize first letter of each word)
    street_address = " ".join([word.capitalize() for word in str(row['Street Address']).lower().split()])
    logger.debug(f"Formatted street address: {street_address}")
    
    # Format suburb (capitalize first letter of each word)
    suburb = " ".join([word.capitalize() for word in str(row['Suburb']).lower().split()])
    logger.debug(f"Formatted suburb: {suburb}")
    
    # Determine price based on property type
    price = ""
    if property_type == 'For Lease':
        if pd.notna(row['Last Rental Price']) and row['Last Rental Price'] != '':
            price = format_price_string(row['Last Rental Price'])
        else:
            price = "Not Disclosed"
    elif property_type == 'For Sale':
        if pd.notna(row['Last Listed Price']) and row['Last Listed Price'] != '':
            price = format_price_string(row['Last Listed Price'])
        else:
            price = "Not Disclosed"
    
    # Format floor area
    floor_area = str(row['Floor Size (m²)']) if pd.notna(row['Floor Size (m²)']) else "Not Disclosed"
    
    # Format car spaces
    car_spaces = str(row['Car']) if pd.notna(row['Car']) else "-"
    
    # Extract property data with key names matching pdf_generator expectations
    property_data = {
        'suburb': str(row['Suburb']),  # Keep original for header
        'suburb_formatted': suburb,    # Formatted for display
        'street address': street_address,  # Note the space in the key name to match pdf_generator
        'floor area': floor_area,      # Space in key name
        'price': price,
        'zoning': str(row['Site Zoning']) if pd.notna(row['Site Zoning']) else "Not Specified",
        'property type': str(row['Property Type']) if pd.notna(row['Property Type']) else "Commercial",
        'car spaces': car_spaces,      # Space in key name
        'comments': str(row["Busi's Comment"]) if pd.notna(row["Busi's Comment"]) else "",
        'image_data': image_data       # Base64 image data from Excel
    }
    
    # Log detailed information about extracted property
    logger.info(f"Extracted property data for {street_address}, {suburb}")
    logger.info(f"Property Type: {property_type}, Price: {price}, Floor Area: {floor_area}")
    logger.info(f"Image data present: {'Yes' if image_data else 'No'}")
    
    # Count of properties processed with images
    if image_data:
        logger.info(f"✅ Property {street_address} has image data")
        logger.info(f"Image data URL starts with: {image_data[:30]}...")
    else:
        logger.info(f"❌ Property {street_address} is missing image data")
    
    return property_data

def format_price_string(price_value):
    """
    Format price values consistently for the report.
    
    Args:
        price_value: The price value to format
        
    Returns:
        str: Formatted price string
    """
    # If price is numeric, format with commas for thousands
    if isinstance(price_value, (int, float)) or (isinstance(price_value, str) and price_value.replace(',', '').replace('.', '').isdigit()):
        # Convert to float first
        if isinstance(price_value, str):
            # Remove currency symbols, commas, etc.
            clean_price = price_value.replace('$', '').replace(',', '')
            try:
                price_num = float(clean_price)
                
                # Format based on value
                if price_num % 1 == 0:  # It's a whole number
                    return f"${int(price_num):,}"
                else:
                    return f"${price_num:,.2f}"
            except ValueError:
                return str(price_value)  # Return as is if conversion fails
        else:
            # It's already a number
            if price_value % 1 == 0:  # It's a whole number
                return f"${int(price_value):,}"
            else:
                return f"${price_value:,.2f}"
    else:
        # If it's already a string like "Offers above $X" or "Not Disclosed"
        return str(price_value)